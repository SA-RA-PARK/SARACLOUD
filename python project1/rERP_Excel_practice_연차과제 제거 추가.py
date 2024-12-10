import pandas as pd
from datetime import datetime
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles import NamedStyle


# 작업할 폴더 경로
folder_path = r"C:\\IACFPYTHON\\MyProject(1)\\work"
file_list = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') and not f.startswith('~$')]

if file_list:
    for file in file_list:
        file_path = os.path.join(folder_path, file)
        df = pd.read_excel(file_path, engine='openpyxl')

        # 필요없는 행 삭제 : D열, M열, U열 필터링
        df_filtered = df[
            (~df.iloc[:,3].astype(str).str.contains('세종대학교|건수', na=False)) &  # D열 필터링
            (~df.iloc[:,12].astype(str).str.contains(r'간접비|_대응|\(대응|\[대응', regex=True, na=False)) &  # M열 필터링 
            (~df.iloc[:,20].astype(str).str.contains('연구산학협력처|산학협력단', na=False))  # U열 필터링
        ]

        # 새 열 추가 작업
        new_column_index = 4  # 0부터 시작하는 인덱스 기준으로 4번째 자리에 삽입
        new_column_name = "과제명(수정)"  # 새 열 이름

        # D열(4번째 열)의 값에서 조건에 따라 처리
        def process_value(value):
            if isinstance(value, str):
                # "["로 시작하면 [~~~] 패턴 제거
                if value.startswith("["):
                    value = re.sub(r'\[.*?\]', '', value).strip()

                # "("로 시작하면 (~~~) 패턴 제거
                if value.startswith("("):
                    value = re.sub(r'\(.*?\)', '', value).strip()

                # "차년도)"로 끝나면 (~~~) 패턴 제거
                if value.endswith("차년도)"):
                    value = re.sub(r'\(.*?\)', '', value).strip()

            return value  # 조건에 맞지 않으면 원래 값 반환

        # 기존 4번째 열의 값을 새 열 데이터로 처리
        new_column_values = df.iloc[:, new_column_index - 1].apply(process_value)

        # 새 열 삽입
        df.insert(new_column_index, new_column_name, new_column_values)

        # 다운로드 날짜 입력하기
        # ①3번째 열(인덱스 2)의 헤드명을 '담당자'에서 '날짜'로 변경
        df_filtered.rename(columns={df_filtered.columns[2]: '날짜'}, inplace=True)

        # ②'날짜' 열의 값(value)을 오늘 날짜로 설정
        df_filtered['날짜'] = pd.to_datetime(datetime.today().strftime('%Y-%m-%d')).date()

        # 파일 저장 경로 설정
        today_date = datetime.today().strftime('%Y%m%d')  # 오늘 날짜 형식 지정
        new_file_path = os.path.join(folder_path, f"수정된_{today_date}_{file}")  # 새로운 파일명 생성

        # 필요한 열만 선택
        selected_columns = df_filtered[['날짜', '과제명(수정)']]  # 날짜와 수정된 과제명 열만 선택

        # 엑셀 파일 저장
        with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
            selected_columns.to_excel(writer, sheet_name=today_date, index=False)  # 날짜를 시트명으로 설정
        print(f"파일 '{file}'을 처리하여 '{new_file_path}'로 저장했습니다.")




        """
        # 오름차순 정렬 : 과제명 > 총연구기간시작일 > 총괄과제총연구비
        # M열(인덱스 12), BB열(인덱스 53), BA열(인덱스 52) 순으로 오름차순 정렬
        df_filtered_sorted = df_filtered.sort_values(by=[df_filtered.columns[12], df_filtered.columns[53], df_filtered.columns[52]], ascending=True)

        
        # 사용할 열만 선택 & 추출
        selected_columns_indices = [2, 3, 12, 15, 16, 20, 21, 52, 53, 54]
        selected_columns = df_filtered.iloc[:, selected_columns_indices]
    

        # 오늘 날짜로 새로운 엑셀 파일 저장
        today_date = datetime.now().strftime('%Y-%m-%d')
        new_file_name = f"{today_date}_{file}"
        new_file_path = os.path.join(folder_path, new_file_name)



        # 엑셀 파일 저장
        with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
            selected_columns.to_excel(writer, sheet_name=today_date, index=False)



            # 열 너비 수동 설정
            workbook = writer.book
            worksheet = workbook[today_date]

            # 열 너비 설정 
            column_widths = {
                'A': 11,   # A열
                'B': 33,   # B열
                'C': 67,   # C열
                'D': 12,   # D열
                'E': 9,    # E열
                'F': 17,   # F열
                'G': 17,   # G열
                'H': 17,   # H열
                'I': 13,   # I열
                'J': 13,   # J열
            }


            # 지정된 열 너비 적용
            for col_letter, col_width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = col_width

            # 첫 번째 행의 높이를 32로 설정
            worksheet.row_dimensions[1].height = 32

            # 첫 번째 행 가운데 정렬
            for cell in worksheet[1]:
                cell.alignment = Alignment(horizontal='center', vertical='center')


            # 글씨 크기 10으로 설정
            font_style = Font(size=10)  # 글씨 크기를 10으로 설정


            # 모든 셀을 '가운데 맞춤'으로 설정
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    # 모든 셀에 글씨 크기 10 적용
                    cell.font = Font(size=10)


            # H열(인덱스 8)의 모든 셀에 '쉼표 스타일' 적용
            for row in worksheet.iter_rows(min_col=8, max_col=8):
                for cell in row:
                    cell.number_format = '#,##0'  # 쉼표 스타일 적용

        print(f"작업 완료: '{new_file_name}'에 데이터가 저장되었습니다.")

else:
    print("작업할 엑셀 파일이 없습니다.")

    
"""