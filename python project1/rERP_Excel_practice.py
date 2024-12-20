import os
import re
import util as ut
from datetime import datetime, timedelta
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl import load_workbook

def clean_value(value):
    if isinstance(value, str):  # 값이 문자열인 경우에만 처리
        if value.startswith("["):  
            value = re.sub(r'\[.*?\]', '', value).strip()
        if value.startswith("("):  
            value = re.sub(r'\(.*?\)', '', value).strip()
        if value.endswith("차년도)"):
            value = re.sub(r'\(.*?차년도\)', '', value).strip()
        if value.endswith("년차)"):
            value = re.sub(r'\(.*?년차\)', '', value).strip()
        if value.endswith("차)"):
            value = re.sub(r'\(.*?차\)', '', value).strip()
    return value

current_path = ut.exedir('py')
folder_path = os.path.join(current_path, "RnD_list\\")
result_folder = os.path.join(folder_path, "result")

if not os.path.exists(result_folder):
    os.makedirs(result_folder)

# 오늘 날짜, 어제 날짜
today_date = datetime.now().strftime('%Y-%m-%d')
yesterday_date = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')

def get_latest_file_in_folder(folder, extension='.xlsx'):
    files = [f for f in os.listdir(folder) if f.endswith(extension) and not f.startswith('~$')]
    if not files:
        return None
    files = sorted(files, key=lambda x: os.path.getmtime(os.path.join(folder, x)), reverse=True)
    return os.path.join(folder, files[0])

# 최초 작업 여부 판단: result 폴더에 이전 결과가 있는지로 판단
previous_result_file = get_latest_file_in_folder(result_folder)

#######################
# 함수 정의: 포맷팅 함수
#######################
def apply_formatting(workbook, today_date):
    worksheet_today = workbook[today_date]
    worksheet_no_duplicates = workbook[f"중복제거({today_date})"]

    worksheet_today.freeze_panes = "A2"
    worksheet_no_duplicates.freeze_panes = "A2"

    column_widths = {
        'A': 11,
        'B': 33,
        'C': 95,
        'D': 13,
        'E': 10,
        'F': 17,
        'G': 17,
        'H': 17,
        'I': 14,
        'J': 14,
    }

    for col_letter, col_width in column_widths.items():
        worksheet_today.column_dimensions[col_letter].width = col_width
        worksheet_no_duplicates.column_dimensions[col_letter].width = col_width

    for row in worksheet_today.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=10)

    for row in worksheet_no_duplicates.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=10)

    worksheet_today.row_dimensions[1].height = 32
    worksheet_no_duplicates.row_dimensions[1].height = 32

    for cell in worksheet_today[1]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name="맑은 고딕", size=10.5, bold=True)
        cell.fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

    for cell in worksheet_no_duplicates[1]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name="맑은 고딕", size=10.5, bold=True)
        cell.fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

    for row in worksheet_today.iter_rows(min_row=2, max_row=worksheet_today.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    for row in worksheet_no_duplicates.iter_rows(min_row=2, max_row=worksheet_no_duplicates.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    for row in worksheet_today.iter_rows(min_row=2, max_row=worksheet_today.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    for row in worksheet_no_duplicates.iter_rows(min_row=2, max_row=worksheet_no_duplicates.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    for row in worksheet_today.iter_rows(min_row=2, max_row=worksheet_today.max_row, min_col=8, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal='right', vertical='center')

    for row in worksheet_no_duplicates.iter_rows(min_row=2, max_row=worksheet_no_duplicates.max_row, min_col=8, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal='right', vertical='center')

    for row in worksheet_today.iter_rows(min_col=8, max_col=8):
        for cell in row:
            cell.number_format = '#,##0'

    for row in worksheet_no_duplicates.iter_rows(min_col=8, max_col=8):
        for cell in row:
            cell.number_format = '#,##0'

    thin_side = Side(border_style="thin", color="000000")  
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row in worksheet_today.iter_rows():
        for cell in row:
            cell.border = thin_border

    for row in worksheet_no_duplicates.iter_rows():
        for cell in row:
            cell.border = thin_border

##########################
# 반복되는 작업 Steps 1~4 함수화
##########################
def initial_processing(df):
    # 1) D,M,U열 필터링
    df_filtered = df[
        (~df.iloc[:, 3].astype(str).str.contains('세종대학교|건수', na=False)) &
        (~df.iloc[:, 12].astype(str).str.contains(r'간접비|_대응|\(대응|\[대응', regex=True, na=False)) &
        (~df.iloc[:, 20].astype(str).str.contains('연구산학협력처|산학협력단', na=False))
    ]

    # 2) 날짜 열 처리
    df_filtered.rename(columns={df_filtered.columns[2]: '날짜'}, inplace=True)
    df_filtered['날짜'] = pd.to_datetime(datetime.today().strftime('%Y-%m-%d')).date()

    # 3) 정렬: M열(12), BB열(53), BA열(52)
    df_filtered_sorted = df_filtered.sort_values(
        by=[df_filtered.columns[12], df_filtered.columns[53], df_filtered.columns[52]],
        ascending=True
    )

    # 4) 사용할 열만 선택 & 추출
    selected_columns_indices = [2, 3, 12, 15, 16, 20, 21, 52, 53, 54]
    selected_columns = df_filtered_sorted.iloc[:, selected_columns_indices]

    return df_filtered_sorted, selected_columns, selected_columns_indices

#############################
# Steps 6~14 수행 함수 정의
#############################
def post_processing(df_filtered_sorted, today_date, yesterday_date, new_file_path):
    # 명시적 복사로 SettingWithCopyWarning 방지
    df_filtered_sorted = df_filtered_sorted.copy()

    # 6~10) 과제명(C열) 관련 정제
    df_filtered_sorted.iloc[:, 2] = df_filtered_sorted.iloc[:, 2].apply(clean_value)
    df_filtered_sorted.iloc[:, 2] = df_filtered_sorted.iloc[:, 2].astype(str).str.replace(r'\(\d+차년도[^\)]*\)', '', regex=True)
    df_filtered_sorted.iloc[:, 2] = df_filtered_sorted.iloc[:, 2].str.replace(r'_\d+차년도$', '', regex=True)
    df_filtered_sorted.iloc[:, 2] = df_filtered_sorted.iloc[:, 2].str.replace(r'\((\d+단계)[-_]\d+차년도\)', r'(\1)', regex=True)
    df_filtered_sorted.iloc[:, 2] = df_filtered_sorted.iloc[:, 2].str.replace(r'(\d+단계)\(\d+차년도\)', r'\1', regex=True)

    # 11) 공백 제거
    df_filtered_sorted.iloc[:, 2] = df_filtered_sorted.iloc[:, 2].str.strip()

    # 12) 재정렬: 연구비(7)-내림차순, 과제명(2)-오름차순, 시작일(8)-오름차순
    df_filtered_sorted = df_filtered_sorted.sort_values(
        by=[df_filtered_sorted.columns[7], df_filtered_sorted.columns[2], df_filtered_sorted.columns[8]],
        ascending=[False, True, True]
    )

    # 필요 열 선택 (여기서는 첫 10개 열 가정)
    selected_columns_cleaned_indices = [0,1,2,3,4,5,6,7,8,9]
    selected_columns_cleaned = df_filtered_sorted.iloc[:, selected_columns_cleaned_indices]

    # 14) 중복 제거: 과제명 컬럼(인덱스 2)
    df_no_duplicates = selected_columns_cleaned.drop_duplicates(subset=[selected_columns_cleaned.columns[2]])


    # 15) 중복제거 후 날짜 컬럼(인덱스 0)에 오늘 날짜로 값 설정
    df_no_duplicates.iloc[:, 0] = pd.to_datetime(datetime.today().strftime('%Y-%m-%d')).date()
    

    with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a') as writer:
        workbook = writer.book

        # "중복제거(어제날짜)" 시트 삭제
        yesterday_sheet_name = f"중복제거({yesterday_date})"
        if yesterday_sheet_name in workbook.sheetnames:
            del workbook[yesterday_sheet_name]

        # "중복제거(오늘날짜)" 시트가 이미 있으면 삭제
        today_sheet_name = f"중복제거({today_date})"
        if today_sheet_name in workbook.sheetnames:
            del workbook[today_sheet_name]

        # 새로 "중복제거(오늘날짜)" 시트 생성
        df_no_duplicates.to_excel(writer, sheet_name=today_sheet_name, index=False)
        apply_formatting(workbook, today_date)
        workbook.save(new_file_path)

########################
# 메인 로직
########################
def rERP_Excel_main() :
    if previous_result_file is None:
        # 최초 실행
        latest_source_file = get_latest_file_in_folder(folder_path)
        if latest_source_file is None:
            print("작업할 엑셀 파일이 없습니다.")
        else:
            df = pd.read_excel(latest_source_file, engine='openpyxl')
            df_filtered_sorted, selected_columns, selected_columns_indices = initial_processing(df)

            new_file_name = f"{today_date}_{os.path.basename(latest_source_file)}"
            new_file_path = os.path.join(result_folder, new_file_name)

            # "오늘날짜" 시트 저장
            with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
                selected_columns.to_excel(writer, sheet_name=today_date, index=False)
                workbook = writer.book
                # 여기서 최소한의 포맷팅 적용 가능(원한다면)
                workbook.save(new_file_path)

            # "중복제거(오늘날짜)" 시트 생성
            # 여기서는 초기 실행이므로 df_filtered_sorted 그대로 사용
            post_processing(df_filtered_sorted, today_date, yesterday_date, new_file_path)

            print(f"최초 작업 완료: '{new_file_path}'에 데이터가 저장되었습니다.")

    else:
        # 최초 실행 이후
        latest_source_file = get_latest_file_in_folder(folder_path)
        if latest_source_file is None:
            print("작업할 엑셀 파일이 없습니다.")
            exit()

        df = pd.read_excel(latest_source_file, engine='openpyxl')
        df_filtered_sorted, selected_columns, selected_columns_indices = initial_processing(df)

        # 어제 날짜로 작업한 result 파일 열기
        yesterday_file_path = None
        for f in os.listdir(result_folder):
            if f.startswith(yesterday_date):
                yesterday_file_path = os.path.join(result_folder, f)
                break

        if yesterday_file_path is None:
            print("어제 날짜 파일을 찾을 수 없습니다. 최초 작업이거나 파일 이름 규칙을 확인하세요.")
            exit()

        # 어제 파일 열어 1번 시트(어제 날짜 시트)에 selected_columns append
        wb = load_workbook(yesterday_file_path)
        if yesterday_date in wb.sheetnames:
            ws = wb[yesterday_date]
            last_row = ws.max_row
            for r_i, row_data in enumerate(selected_columns.values, start=last_row+1):
                for c_i, val in enumerate(row_data, start=1):
                    ws.cell(row=r_i, column=c_i, value=val)

            # 시트 이름을 오늘날짜로 변경
            wb[yesterday_date].title = today_date
            updated_file_path = os.path.join(result_folder, f"{today_date}_{os.path.basename(latest_source_file)}")
            wb.save(updated_file_path)
            wb.close()

            # 7) 오늘날짜로 새로 저장한 시트를 읽고 post_processing 진행
            df_for_post = pd.read_excel(updated_file_path, sheet_name=today_date, engine='openpyxl')
            df_filtered_sorted_post = df_for_post.copy()

            # post_processing 호출 시 df_filtered_sorted_post, today_date, yesterday_date, new_file_path 사용
            # new_file_path는 updated_file_path와 동일하게 사용 가능
            new_file_path = updated_file_path
            post_processing(df_filtered_sorted_post, today_date, yesterday_date, new_file_path)

            print(f"일일 작업 완료: '{updated_file_path}'에 오늘날짜 시트 갱신 및 중복제거 시트 저장 완료.")
        else:
            print("어제 날짜 시트를 찾을 수 없습니다. 파일 구조를 확인하세요.")
